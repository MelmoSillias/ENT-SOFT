#!/usr/bin/env python3
"""Generate ENT-SOFT SQL seed from Planning + Survey Excel files."""
from __future__ import annotations

import json
import re
import uuid
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

PLANNING = Path(r"C:\Users\DELL\Desktop\ENT\PLaning ENT for 22sites on august.xlsx")
ATEL = Path(r"C:\Users\DELL\Desktop\ENT\suivi ATEL.xlsx")
OUT = Path(r"C:\Users\DELL\Desktop\Programmation\Web\ENT-SOFT\api\migrations\seeds\seed_telecel_planning.sql")

NS = uuid.UUID("a1b2c3d4-e5f6-7890-abcd-ef1234567890")

def uid(key: str) -> str:
    return str(uuid.uuid5(NS, key))

def bin_uuid(key: str) -> str:
    """SQLite BLOB literal (Doctrine uuid type stores 16-byte binary)."""
    return f"X'{uid(key).replace('-', '')}'"

def esc(s) -> str:
    if s is None:
        return "NULL"
    return "'" + str(s).replace("\\", "\\\\").replace("'", "''") + "'"

def jesc(obj) -> str:
    return esc(json.dumps(obj, ensure_ascii=False))

def excel_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(v))).date().isoformat()
    except Exception:
        return None

def month_fr(iso):
    if not iso:
        return "?"
    months = ["", "janvier", "fevrier", "mars", "avril", "mai", "juin",
              "juillet", "aout", "septembre", "octobre", "novembre", "decembre"]
    y, m, _ = iso.split("-")
    return f"{months[int(m)]} {y}"

def norm_name(n):
    if not n:
        return ""
    n = re.sub(r"\s+", " ", str(n)).strip()
    low = n.lower()
    if "moussa" in low and "ouatt" in low:
        return "Moussa Ouattara"
    if "sory" in low and "dram" in low:
        return "Sory Drame"
    if "alasane" in low or "alassane" in low:
        return "Alasane Sanogo"
    return n

def map_status(*raws):
    rank = {"completed": 4, "in_progress": 3, "blocked": 2, "pending": 1}
    best = "pending"
    for r in raws:
        if not r:
            continue
        s = str(r).strip().lower()
        st = "pending"
        if "done" in s or "termin" in s or s == "completed":
            st = "completed"
        elif "inprogress" in s or "in progress" in s or "en cours" in s or s == "in_progress":
            st = "in_progress"
        elif "probleme" in s or "probleme" in s or "block" in s or s == "blocked":
            st = "blocked"
        if rank[st] > rank[best]:
            best = st
    return best

def cell(row, idx):
    if idx >= len(row):
        return None
    v = row[idx]
    return v.strip() if isinstance(v, str) else v

sites = {}

def add_site(code, title=None):
    code = str(code).strip().upper()
    if not code:
        return
    title = (str(title).strip() if title else "")
    if code not in sites:
        sites[code] = title if title else code
    elif title and (not sites[code] or sites[code] == code):
        sites[code] = title
    if code == "BMK1035":
        sites[code] = "Niamakoro"
    if code == "BMK1157":
        sites[code] = "C"
    if code == "KLK2182":
        sites[code] = "Kati_Sud_Ecole"

BASE = [
("BMK1003","Banconi"),("BMK1005","Grand Marche"),("BMK1007","Djicoroni_para_EDM"),
("BMK1008","Coura Bolibana"),("BMK1011","Djikoroni_para"),("BMK1015","Hamdallaye ACI_FR3"),
("BMK1016","Hippodrome"),("BMK1019","Kalaban_ACI"),("BMK1020","Kalaban Coro Sud"),
("BMK1022","Kalaban_Coura-Sud"),("BMK1024","Korofina Nord"),("BMK1025","1008 Logements"),
("BMK1028","Moribabougou 2"),("BMK1029","Magnambougou Nord"),("BMK1034","Ngolonina"),
("BMK1035","Niamakoro"),("BMK1037","Tomikorobougou_GMS"),("BMK1039","Petit_Paris"),
("BMK1040","Point_G"),("BMK1041","Quartier_du_Fleuve"),("BMK1043","Doumanzana Est"),
("BMK1044","Sabalibougou"),("BMK1045","Titibougou 2"),("BMK1046","Sirakoro"),
("BMK1049","Sotuba"),("BMK1050","Stade 26 Mars"),("BMK1051","Taliko"),
("BMK1052","Tiebani"),("BMK1053","Boulkassoumbougou"),("BMK1055","Missabougou"),
("BMK1056","Bougouba"),("BMK1058","Baco ACI"),("BMK1060","Sokorodji"),
("BMK1061","Sebenikoro SEMA"),("BMK1063","Boulkassoumbougou Nord"),("BMK1069","Djikoroni Para"),
("BMK1070","Zone Industrielle"),("BMK1071","Titibougou"),("BMK1072","Hippodrome_Nord"),
("BMK1073","Koulouba_Village"),("BMK1074","Koyambougou"),("BMK1077","Niamakoro Diallabougou"),
("BMK1078","Niamakoro_Sud"),("BMK1080","Senou 1"),("BMK1082","Same"),
("BMK1083","Sotuba IER"),("BMK1085","Banconi Ouest"),("BMK1086","Senou 2"),
("BMK1087","Missira"),("BMK1091","Sebenincoro Marche"),("BMK1092","Djikoroni Coura"),
("BMK1094","Hamdalaye_Marche"),("BMK1096","Djikoroni_Para_Sud"),("BMK1097","Moribabougou"),
("BMK1098","Doumazana Sud"),("BMK1099","Dialakorodji Nord"),("BMK1101","Banconi Nord"),
("BMK1103","Sabalibougou EST"),("BMK1108","Sogoniko Gare"),("BMK1109","Faladie Ouest"),
("BMK1113","Nafadji"),("BMK1114","Niamakoro ouest"),("BMK1115","Kabala2"),
("BMK1116","Sabalibougou_Terminus"),("BMK1119","Ngolobougou"),("BMK1120","Kalaban Coro Nord"),
("BMK1125","Djicoroni_Para_Usine"),("BMK1126","Sebenikoro Champ"),("BMK1128","ACI2000_Nord"),
("BMK1131","Maison des jeunes"),("BMK1135","Mamaribougou"),("BMK1138","Kanadjiguila 2"),
("BMK1139","Lafiabougou Terrain"),("BMK1140","Yirimadjo Campement"),("BMK1141","Baco EST"),
("BMK1142","Daoudabougou Ouest"),("BMK1147","Bagadadji_Nord"),("BMK1149","Doumazana OUEST"),
("BMK1154","Hampate Bah"),("BMK1155","Palais de la Culture"),("BMK1157","C"),
("BMK1158","Samaya_Village"),("BMK1159","Gana"),("BMK1162","Sebenikoro_Cite"),
("BMK1172","Kabala"),("BMK1174","Yiramadjo_wara"),("BMK1176","Missabougou_Sirafara"),
("BMK1177","Dianeguela"),("BMK1178","Plazio_Orange"),
("KLK1042","Dialakorodji"),("KLK2180","Koulikoro_Huicoma"),("KLK2181","Kati_Malibougou"),
("KLK2182","Kati_Sud_Ecole"),("KLK2185","Kassela"),("KLK2186","Koulikoroba 2"),
("KLK2190","Baguineda"),("KLK2191","Farabana"),("KLK2221","Didieni"),
("KYS2231","Tabakoto"),("KYS3231","Tabakoto"),("KYS3232","DJIDJAN MINE"),
("KYS3233","DJIDJAN VILLE"),("KYS3234","Kenieba_FO"),("KYS3323","KENIEBA LAFIABOUGOU"),
("SEG4252","Diola"),("SKS5206","Gouvernorat Sikasso"),("SKS5213","KOUTIALA_NORD"),
("SKS5217","Koutiala_Cimetiere"),("SKS5270","Kolondieba"),("SKS5280","Niena"),
("SKS5289","Sikasso ATTBOUGOU"),
]
for c,t in BASE:
    add_site(c,t)

for n in ["1010","1014","1021","1026","1032","1038","1048","1057","1064","1066","1067","1079","1088","1105","1106","1111","1112","1118","1121"]:
    add_site(f"BMK{n}", f"Site BMK{n}")

# Regional: San/Bla=SEG (Segou), Sevare/Mopti=MPT (Mopti)
# User: Sevare cocam=6237, Tahikiri=6238. Carrefou Excel also 6238 -> Carrefou MPT6244
PLAN13 = [
("SEG4244","san -Santoro"),("SEG4245","san -Lafiabougou"),("SEG4246","san -Bagadadji"),
("SEG4247","san -Medine"),("KLK2182","Kati_Sud_Ecole"),("SEG4248","Bla -Carrafour"),
("MPT6244","Sevare-Carrefou"),("MPT6240","Sevare-Banguetaba"),("MPT6241","Sevare-sokoura"),
("MPT6242","Sevare-sarema"),("MPT6237","Sevare-Sevare cocam"),("MPT6238","Mopti-Tahikiri"),
("MPT6239","Mopti-Medine"),("MPT6236","Mopti-Mossinkore"),
]
for c,t in PLAN13:
    add_site(c,t)

PLAN13_MAP = {
4244:"SEG4244",4245:"SEG4245",4246:"SEG4246",4247:"SEG4247",2182:"KLK2182",4248:"SEG4248",
6238:"MPT6244",6240:"MPT6240",6241:"MPT6241",6242:"MPT6242",6237:"MPT6237",6239:"MPT6239",6236:"MPT6236",
}

wb_p = load_workbook(PLANNING, data_only=True)
wb_a = load_workbook(ATEL, data_only=True)

employees = {"Moussa Ouattara":"50610050","Sory Drame":"93376133","Alasane Sanogo":"52228385"}

ws1 = wb_p["Feuil1"]
feuil1_rows = []
current_lot = None
for row in ws1.iter_rows(min_row=2, max_row=44, values_only=True):
    lot = cell(row, 0)
    if lot:
        current_lot = str(lot).strip()
    code = cell(row, 3)
    if not code:
        continue
    code = str(code).strip().upper()
    add_site(code, cell(row, 2))
    tech = norm_name(str(cell(row, 19) or ""))
    phone = str(cell(row, 20)).strip() if cell(row, 20) else None
    if tech and phone:
        employees[tech] = phone
    vals = {
        "ville": cell(row,1), "config_2100_actuel": cell(row,4), "config_2100_planifie": cell(row,5),
        "configuration": cell(row,6), "swap_rru_2100": cell(row,7), "swap_rru_900": cell(row,8),
        "swap_rru_1800": cell(row,9), "swap_antenne": cell(row,10),
        "nombre_rru_a_demonter": cell(row,11), "nombre_rru_a_monter": cell(row,12),
        "nombre_antenne_a_monter": cell(row,13), "nombre_jumper": cell(row,14),
        "longitude": cell(row,15), "latitude": cell(row,16),
        "start_date": excel_date(cell(row,17)), "end_date": excel_date(cell(row,18)),
        "team_name": tech or None, "team_contact": phone,
    }
    feuil1_rows.append({"lot": current_lot or "LOT1", "code": code, "vals": vals, "status": "pending", "tech": tech or None})

ws2 = wb_p["Feuil2"]
feuil2_t1 = []
for row in ws2.iter_rows(min_row=3, max_row=16, values_only=True):
    name = cell(row, 0)
    sid = cell(row, 1)
    if not sid or not name:
        continue
    try:
        sid_i = int(float(sid))
    except Exception:
        continue
    name_s = str(name).strip()
    if sid_i == 6237 and "tahikiri" in name_s.lower():
        code = "MPT6238"
    elif sid_i == 6238 and "carrefou" in name_s.lower():
        code = "MPT6244"
    else:
        code = PLAN13_MAP.get(sid_i)
        if not code:
            continue
    add_site(code, name_s if code != "KLK2182" else "Kati_Sud_Ecole")
    tech = norm_name(str(cell(row,5) or ""))
    feuil2_t1.append({
        "code": code, "start": excel_date(cell(row,2)), "end": excel_date(cell(row,3)),
        "tech": tech, "status": map_status(str(cell(row,6) or "")),
        "vals": {"nbre_pv": cell(row,4), "start_date": excel_date(cell(row,2)),
                 "end_date": excel_date(cell(row,3)), "status_raw": cell(row,6)},
    })

feuil2_t2 = {}
for row in ws2.iter_rows(min_row=19, max_row=40, values_only=True):
    code = cell(row, 1)
    if not code or str(code).upper() == "TOTAL":
        continue
    code = str(code).strip().upper()
    if not re.match(r"^[A-Z]{2,4}\d+$", code):
        continue
    add_site(code, cell(row,2))
    feuil2_t2[code] = {
        "status": map_status(str(cell(row,10) or "")),
        "tech": norm_name(str(cell(row,8) or "")),
        "phone": str(cell(row,9)).strip() if cell(row,9) else None,
    }

ws4 = wb_p["Feuil4"]
feuil4_rows = []
starts, ends = [], []
rank = {"completed":4,"in_progress":3,"blocked":2,"pending":1}
for row in ws4.iter_rows(min_row=3, max_row=24, values_only=True):
    code = cell(row,1)
    if not code or str(code).upper()=="TOTAL":
        continue
    code = str(code).strip().upper()
    add_site(code, cell(row,2))
    st4 = map_status(str(cell(row,10) or ""))
    st2 = feuil2_t2.get(code,{}).get("status","pending")
    status = st4 if rank[st4] >= rank[st2] else st2
    tech = norm_name(str(cell(row,8) or "")) or feuil2_t2.get(code,{}).get("tech","")
    phone = (str(cell(row,9)).strip() if cell(row,9) else None) or feuil2_t2.get(code,{}).get("phone")
    if tech and phone:
        employees[tech] = phone
    sd, ed = excel_date(cell(row,6)), excel_date(cell(row,7))
    if sd: starts.append(sd)
    if ed: ends.append(ed)
    feuil4_rows.append({"code":code,"status":status,"tech":tech,"vals":{
        "type_pylone":cell(row,3),"nombre_panneaux_595w":cell(row,4),
        "nombre_regulateurs_victron":cell(row,5),"start_date":sd,"end_date":ed,
        "comment":cell(row,11),"status_raw":cell(row,10),
    }})

proj_pv_start = min(starts) if starts else None
proj_pv_end = max(ends) if ends else None
proj_pv_title = f"Installation de pv {month_fr(proj_pv_start)} - {month_fr(proj_pv_end)}"
t1_starts = [r["start"] for r in feuil2_t1 if r["start"]]
t1_ends = [r["end"] for r in feuil2_t1 if r["end"]]
t1_start = min(t1_starts) if t1_starts else None
t1_end = max(t1_ends) if t1_ends else None
proj_13_title = f"Installation de pv {month_fr(t1_start)} - {month_fr(t1_end)}"

ws_s = wb_a["Survey"]
survey_by_code = {}
for row in ws_s.iter_rows(min_row=1, max_row=51, values_only=True):
    sid = cell(row,1)
    if sid is None:
        continue
    sid_s = str(sid).strip().lstrip(":").strip()
    if not re.match(r"^\d{3,5}$", sid_s):
        continue
    code = "KLK2190" if sid_s=="2190" else f"BMK{sid_s}"
    add_site(code, sites.get(code, f"Site {code}"))
    survey_by_code[code] = {"code":code,"status":"pending","vals":{
        "meplat":cell(row,2),"vert_jaune":cell(row,3),"cable_nu":cell(row,4),
        "barrette_de_terre":cell(row,5),"puit_de_terre":cell(row,6),"remarques":cell(row,7),
        "technicien": None,
    }}
survey_rows = list(survey_by_code.values())

# ---------- ATEL solaire : site Propose + SITE installer + Feuil1 → 1 projet ----------
def parse_atel_solar_sheet(ws, has_stat=True):
    rows_out = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 80, values_only=True):
        code = cell(row, 1)
        name = cell(row, 2)
        if code is None and name:
            # Dialakorodji sans code dans site Propose → KLK1042
            if str(name).strip().lower().startswith("dialakorodji"):
                code = "KLK1042"
            else:
                continue
        if code is None:
            continue
        code_s = str(code).strip().upper()
        if code_s in ("SITE ID", "N", "N°", "TOTAL") or not re.match(r"^[A-Z]{2,4}\d+$", code_s):
            continue
        add_site(code_s, name)
        raw_stat = cell(row, 6) if has_stat else None
        rows_out.append({
            "code": code_s,
            "status": map_status(str(raw_stat or "")),
            "vals": {
                "deploiement_solaire": cell(row, 3),
                "nombre_panneaux": cell(row, 4),
                "nombre_regulateurs_victron": cell(row, 5),
                "status_raw": raw_stat,
                "technicien": None,
            },
        })
    return rows_out

atel_merged = {}  # code -> {status, vals, sources}
for src_name, parsed in [
    ("site Propose", parse_atel_solar_sheet(wb_a["site Propose"], True)),
    ("SITE installer", parse_atel_solar_sheet(wb_a["SITE installer"], True)),
    ("Feuil1", parse_atel_solar_sheet(wb_a["Feuil1"], False)),
]:
    for r in parsed:
        code = r["code"]
        if code not in atel_merged:
            atel_merged[code] = {
                "code": code,
                "status": r["status"],
                "vals": dict(r["vals"]),
                "sources": [src_name],
            }
        else:
            cur = atel_merged[code]
            cur["sources"].append(src_name)
            # highest status wins
            if rank.get(r["status"], 0) > rank.get(cur["status"], 0):
                cur["status"] = r["status"]
            # fill empty complementary values; prefer non-empty newer values
            for k, v in r["vals"].items():
                if k == "technicien":
                    cur["vals"]["technicien"] = None
                    continue
                if k == "status_raw":
                    # keep raw of highest status source when possible
                    if rank.get(r["status"], 0) >= rank.get(cur["status"], 0) and v not in (None, ""):
                        cur["vals"]["status_raw"] = v
                    elif cur["vals"].get("status_raw") in (None, "") and v not in (None, ""):
                        cur["vals"]["status_raw"] = v
                    continue
                if v not in (None, ""):
                    if cur["vals"].get(k) in (None, ""):
                        cur["vals"][k] = v
                    elif src_name == "SITE installer":
                        # installer is execution truth for quantities when present
                        cur["vals"][k] = v
            cur["vals"]["technicien"] = None

atel_rows = sorted(atel_merged.values(), key=lambda x: x["code"])
for r in atel_rows:
    r["vals"]["sources"] = ", ".join(sorted(set(r["sources"])))
    r["vals"]["technicien"] = None

now = "2026-09-03 12:00:00"
lines = []
lines.append("-- Seed ENT-SOFT : Telecel, techniciens, sites, projets Planning + Survey")
lines.append("-- Genere pour SQLite (UUID en BLOB hex)")
lines.append("-- Prerequis : migration Version20260903122230")
lines.append("")
client_bin = bin_uuid("client:telecel")
lines.append("-- === Client Telecel ===")
lines.append(f"INSERT INTO clients (id, code, title, description, created_at, updated_at, is_enabled) VALUES ({client_bin}, 'CLI-TELECEL', 'Telecel', 'Operateur telecoms — projets ENT', '{now}', '{now}', 1);")
lines.append(f"INSERT INTO client_contacts (id, client_id, name, phone, created_at) VALUES ({bin_uuid('contact:telecel:ops')}, {client_bin}, 'Contact operations Telecel', '00000000', '{now}');")
lines.append("")
lines.append("-- === Techniciens ===")
for name, phone in employees.items():
    email = re.sub(r"[^a-z0-9]+", ".", name.lower()).strip(".") + "@ent.local"
    lines.append(f"INSERT INTO employees (id, name, email, phone, address, \"function\", user_id, created_at, updated_at, is_enabled) VALUES ({bin_uuid('emp:'+name)}, {esc(name)}, {esc(email)}, {esc(phone or '00000000')}, NULL, 'technicien', NULL, '{now}', '{now}', 1);")
lines.append("")
lines.append(f"-- === Sites ({len(sites)}) ===")
for code in sorted(sites.keys()):
    lines.append(f"INSERT INTO sites (id, code, title, description, client_id, created_at, updated_at, is_enabled) VALUES ({bin_uuid('site:'+code)}, {esc(code)}, {esc(sites[code])}, NULL, {client_bin}, '{now}', '{now}', 1);")
lines.append("")

def insert_project(key, code, title, status, obj, d1, d2, infos):
    lines.append(f"INSERT INTO projects (id, code, title, object, date_debut, date_fin, status, budget, client_id, sites_informations, created_at, updated_at, is_enabled) VALUES ({bin_uuid(key)}, {esc(code)}, {esc(title)}, {esc(obj)}, {esc(d1) if d1 else 'NULL'}, {esc(d2) if d2 else 'NULL'}, {esc(status)}, 0, {client_bin}, {jesc(infos)}, '{now}', '{now}', 1);")

def insert_lot(proj_key, lot_code, lot_title):
    lines.append(f"INSERT INTO project_lots (id, project_id, code, title, created_at, updated_at) VALUES ({bin_uuid('lot:'+proj_key+':'+lot_code)}, {bin_uuid(proj_key)}, {esc(lot_code)}, {esc(lot_title)}, '{now}', '{now}');")

def insert_ps(proj_key, site_code, status, vals, tech=None, lot_code=None, idx=0):
    tech_sql = bin_uuid("emp:"+tech) if tech and tech in employees else "NULL"
    emp_json = jesc([uid("emp:"+tech)] if tech and tech in employees else [])
    lot_sql = bin_uuid("lot:"+proj_key+":"+lot_code) if lot_code else "NULL"
    lines.append(f"INSERT INTO project_sites (id, project_id, site_id, lot_id, technician_id, status, date_added, informations_values, employee_ids, created_at, updated_at) VALUES ({bin_uuid(f'ps:{proj_key}:{site_code}:{idx}')}, {bin_uuid(proj_key)}, {bin_uuid('site:'+site_code)}, {lot_sql}, {tech_sql}, {esc(status)}, '{now}', {jesc(vals)}, {emp_json}, '{now}', '{now}');")

lines.append("-- === Projet Telecel (Feuil1) pending + lots ===")
infos1 = [{"key":k,"label":l} for k,l in [
("ville","Ville"),("config_2100_actuel","CONFIG 2100 ACTUEL"),("config_2100_planifie","CONFIG 2100 PLANIFIE"),
("configuration","CONFIGURATION"),("swap_rru_2100","Swap RRU 2100"),("swap_rru_900","Swap RRU 900"),
("swap_rru_1800","Swap RRU 1800"),("swap_antenne","SWAP ANTENNE"),("nombre_rru_a_demonter","Nombre RRU a demonter"),
("nombre_rru_a_monter","Nombre RRU a monter"),("nombre_antenne_a_monter","Nombre d antenne a monter"),
("nombre_jumper","NOMBRE DE Jumper"),("longitude","Longitude"),("latitude","Latitude"),
("start_date","Start Date"),("end_date","End Date"),("team_name","Team Name"),("team_contact","Team Contact")]]
insert_project("proj:telecel-swap","PRJ-TELECEL-SWAP","Projet Telecel","pending","Swap/upgrade RRU — Planning Feuil1 (LOT1+LOT2)",None,None,infos1)
insert_lot("proj:telecel-swap","LOT1","Lot 1")
insert_lot("proj:telecel-swap","LOT2","Lot 2")
for i,r in enumerate(feuil1_rows):
    insert_ps("proj:telecel-swap", r["code"], r["status"], r["vals"], tech=r.get("tech"), lot_code=r["lot"], idx=i)
lines.append("")

lines.append("-- === Installation PV 13 sites (Feuil2 T1) completed ===")
infos13 = [{"key":"nbre_pv","label":"Nbre de PV"},{"key":"start_date","label":"Start Date"},{"key":"end_date","label":"End Date"},{"key":"status_raw","label":"Status source"}]
insert_project("proj:pv-13","PRJ-PV-13SITES",proj_13_title,"completed","Planning Feuil2 tableau 1 (13 sites) — closed",t1_start,t1_end,infos13)
for i,r in enumerate(feuil2_t1):
    insert_ps("proj:pv-13", r["code"], r["status"], r["vals"], tech=r["tech"] or None, idx=i)
lines.append("")

lines.append("-- === Installation PV 22 sites (Feuil4) active ===")
infos22 = [{"key":k,"label":l} for k,l in [
("type_pylone","Type de pylone"),("nombre_panneaux_595w","Nombre de panneaux 595 watt"),
("nombre_regulateurs_victron","Nombre regulateurs Victron"),("start_date","Start Date"),
("end_date","End Date"),("comment","Commentaire"),("status_raw","Status source")]]
insert_project("proj:pv-22","PRJ-PV-22SITES",proj_pv_title,"active","Planning Feuil4 (maj Feuil2 T2) — in progress",proj_pv_start,proj_pv_end,infos22)
for i,r in enumerate(feuil4_rows):
    insert_ps("proj:pv-22", r["code"], r["status"], r["vals"], tech=r["tech"] or None, idx=i)
lines.append("")

lines.append("-- === Survey mise a la terre (suivi ATEL) ===")
infos_s = [{"key":k,"label":l} for k,l in [
("meplat","Meplat"),("vert_jaune","Vert/Jaune"),("cable_nu","Cable Nu"),
("barrette_de_terre","Barrette de terre"),("puit_de_terre","Puit de terre"),("remarques","Remarques"),
("technicien","Technicien")]]
insert_project("proj:survey","PRJ-SURVEY-ATEL","Survey mise a la terre","active","suivi ATEL feuille Survey (Feuil2 doublon non seede)",None,None,infos_s)
for i,r in enumerate(survey_rows):
    insert_ps("proj:survey", r["code"], r["status"], r["vals"], idx=i)
lines.append("")

lines.append("-- === Deploiement solaire ATEL (site Propose + SITE installer + Feuil1) ===")
infos_atel = [{"key":k,"label":l} for k,l in [
("deploiement_solaire","Deploiement solaire"),
("nombre_panneaux","Nombre panneaux"),
("nombre_regulateurs_victron","Besoin regulateurs VICTRON"),
("status_raw","Status source"),
("technicien","Technicien"),
("sources","Sources feuilles")]]
# project status = active if any site not completed, else completed
atel_proj_status = "completed"
for r in atel_rows:
    if r["status"] != "completed":
        atel_proj_status = "active"
        break
insert_project(
    "proj:atel-solaire",
    "PRJ-ATEL-SOLAIRE",
    "Deploiement solaire ATEL",
    atel_proj_status,
    "Fusion suivi ATEL : site Propose + SITE installer + Feuil1 (technicien vide)",
    None, None, infos_atel,
)
for i, r in enumerate(atel_rows):
    # technicien volontairement vide (NULL) — champ present dans informationsValues
    insert_ps("proj:atel-solaire", r["code"], r["status"], r["vals"], tech=None, idx=i)

OUT.write_text("\n".join(lines)+"\n", encoding="utf-8")
# fix accents
text = OUT.read_text(encoding="utf-8")
text = text.replace("aout", "ao\u00fbt").replace("mise a la terre", "mise \u00e0 la terre")
text = text.replace("Deploiement", "D\u00e9ploiement").replace("Operateur", "Op\u00e9rateur")
OUT.write_text(text, encoding="utf-8")
print(f"Wrote {OUT} ({OUT.stat().st_size} bytes)")
print(f"Sites={len(sites)} Feuil1={len(feuil1_rows)} PV13={len(feuil2_t1)} PV22={len(feuil4_rows)} Survey={len(survey_rows)} ATEL={len(atel_rows)}")
print(f"Titles: {proj_13_title} | {proj_pv_title}")
print("ATEL project status:", atel_proj_status)
print("Emps:", employees)
